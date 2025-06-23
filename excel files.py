import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

wb = Workbook()
sheet = wb.active
sheet.title = "Present"

sheet["A1"] = "Name"
sheet["A2"] = "Age"
sheet.append(["Madhan", 25])
sheet.append(["Maddy", 25])
sheet.append(["Jose", 25])

wb.save(r"C:\HEXAWARE\HEXAWARE-CODINGPRACTISE\hexaware.xlsx")

wb = load_workbook(r"C:\HEXAWARE\HEXAWARE-CODINGPRACTISE\hexaware.xlsx")
sheet = wb.active
for row in sheet.iter_rows(values_only = True):
    print(row)


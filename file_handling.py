with open("C:\\HEXAWARE\\HEXAWARE-CODINGPRACTISE\\GITHUB REPOS.txt","w") as file:
    file.write("hello hexaware")

with open("C:\\HEXAWARE\\HEXAWARE-CODINGPRACTISE\\GITHUB REPOS.txt","r") as file:
    content = file.read()
    print(content)



with open ("C:\\HEXAWARE\\HEXAWARE-CODINGPRACTISE\\GITHUB REPOS.txt","a") as file:
    file.write("adding the present")




from openpyxl import Workbook


wb = Workbook()
ws = wb.active

for i in range(1, 1000001):
    ws.append([f"Row {i}", f"Value {i}"])

wb.save(r"C:\HEXAWARE\HEXAWARE-CODINGPRACTISE\hexaware.xlsx")



from openpyxl import load_workbook

wb = load_workbook(r"C:\HEXAWARE\HEXAWARE-CODINGPRACTISE\hexaware.xlsx", read_only=True)
ws = wb.active


for row in ws.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True):
    print(row)